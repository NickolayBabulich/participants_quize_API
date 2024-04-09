from django.contrib import admin, messages
from django.http import HttpResponse
from participants.models import Participants, TestSending, Log, PhoneVerify, Winners
from dotenv import load_dotenv
import requests
import datetime
import csv
import openpyxl
import os


load_dotenv()


@admin.register(Winners)
class WinnersAdmin(admin.ModelAdmin):
    list_display = [field.name for field in Winners._meta.get_fields()]


@admin.register(Participants)
class UserRegisterAdmin(admin.ModelAdmin):
    list_display = [field.name for field in Participants._meta.get_fields()]
    search_fields = ('first_name', 'second_name', 'phone')
    actions = ['data_send', 'export_to_csv', 'export_to_excel', 'send_to_sms', 'delete_selected']
    actions_selection_counter = True

    def get_action_choices(self, request, **kwargs):
        choices = super(UserRegisterAdmin, self).get_action_choices(request)
        index = next((i for i, choice in enumerate(choices) if choice[0] == ''), None)
        if index is not None:
            choices[index] = ('', 'Выберите действие:')
        return choices

    @admin.action(description="Отправить данные")
    def data_send(self, request, queryset):
        sent_successfully = False
        for object in queryset:
            if not object.status:
                url = os.getenv('DESTINATION_URL')
                data = {
                    'first_name': object.first_name,
                    'second_name': object.second_name,
                    'patronymic': object.patronymic,
                    'birthday': object.birthday,
                    'district_id': object.district_id,
                    'address': object.address,
                    'email': object.email,
                    'phone': object.phone
                }
                try:
                    response = requests.post(url, data=data)
                    response.raise_for_status()
                    object.status = True
                    object.timestamp = datetime.datetime.now()
                    object.save()
                    Log.objects.create(user_id=object.pk, email=object.email, status='accepted',
                                       server_response=response.status_code, timestamp=datetime.datetime.now())
                    sent_successfully = True
                except requests.exceptions.RequestException as error:
                    Log.objects.create(user_id=object.pk, email=object.email, status='rejected',
                                       server_response=response.status_code, timestamp=datetime.datetime.now())
        if sent_successfully:
            messages.success(request, "Данные успешно отправлены")
        else:
            if object.status:
                messages.error(request, "Выбранные данные уже отправлены")
            else:
                messages.error(request, "Ошибка отправки данных")

    @admin.action(description="Выгрузить в CSV")
    def export_to_csv(self, request, queryset):
        try:
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="participants.csv"'

            writer = csv.writer(response, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(
                ['first_name', 'second_name', 'patronymic', 'birthday', 'district_id', 'address', 'email', 'phone'])

            for object in queryset:
                writer.writerow(
                    [object.first_name, object.second_name, object.patronymic, object.birthday, object.district_id,
                     object.address, object.email, object.phone])
            return response
        except requests.exceptions.RequestException as error:
            messages.error(request, "Ошибка выгрузки данных в формате CSV")

    @admin.action(description="Выгрузить в Excel")
    def export_to_excel(self, request, queryset):
        try:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="participants.xlsx"'

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Participants'

            columns = ['first_name', 'second_name', 'patronymic', 'birthday', 'district_id', 'address', 'email',
                       'phone']
            for col_num, col_title in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=col_title)

            for row_num, object in enumerate(queryset, 2):
                row = [object.first_name, object.second_name, object.patronymic, object.birthday, object.district_id,
                       object.address, object.email, object.phone]
                for col_num, col_title in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=col_title)

            wb.save(response)
            return response
        except requests.exceptions.RequestException as error:
            messages.error(request, "Ошибка выгрузки данных в формате Excel")

    @admin.action(description='Отправить SMS')
    def send_to_sms(self, request, queryset):
        sent_successfully = False
        for object in queryset:
            if not object.time_sms_send:
                sending = (int(object.phone), f"проверка рассылки")
                if sending:
                    object.time_sms_send = datetime.datetime.now()
                    object.save()
                    sent_successfully = True
            else:
                print('sms already sent')
        if sent_successfully:
            messages.success(request, "SMS успешно отправлены")


@admin.register(TestSending)
class TestAdmin(admin.ModelAdmin):
    list_display = [field.name for field in TestSending._meta.get_fields()]
    search_fields = ('pid', 'title')


@admin.register(Log)
class LogAdmin(admin.ModelAdmin):
    list_display = [field.name for field in Log._meta.get_fields()]
    list_filter = ('status', 'server_response')


@admin.register(PhoneVerify)
class PhoneVerifyAdmin(admin.ModelAdmin):
    list_display = ('phone', 'code')
